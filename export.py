import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

MONTHS_RU = ['Январь','Февраль','Март','Апрель','Май','Июнь',
             'Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']

CAT_LABELS = {'team':'Команда', 'ads':'Реклама', 'tools':'Сервисы',
              'tax':'Налоги', 'content':'Контент', 'other':'Другое'}
TAG_LABELS = {'client':'Клиент', 'neuro':'Нейроспринт',
              'consult':'Консультация', 'other':'Другое'}
TYPE_LABELS = {'tracking':'Трекинг', 'master':'Мастер-группа', 'project':'Проект'}
STATUS_LABELS = {'active':'В работе', 'new':'Новый', 'done':'Завершён'}

HEADER_FILL = PatternFill("solid", fgColor="2563EB")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)


def _style_header_row(ws, row_num, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def build_excel(incomes, expenses, clients, streams, year, month) -> bytes:
    wb = Workbook()
    period = f"{MONTHS_RU[month]} {year}"

    # --- Финансы ---
    ws = wb.active
    ws.title = "Финансы"
    ws.append([f"Финансы — {period}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    # Доходы
    ws.append(["ДОХОДЫ"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.append(["Дата", "Источник", "Описание", "Тег", "Сумма, ₽"])
    _style_header_row(ws, ws.max_row, 5)
    total_inc = 0
    for i in incomes:
        ws.append([i.date, i.source, i.desc, TAG_LABELS.get(i.tag, i.tag), i.amount])
        total_inc += i.amount
    ws.append(["", "", "", "ИТОГО", total_inc])
    ws.cell(row=ws.max_row, column=4).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=5).font = Font(bold=True)
    ws.append([])

    # Расходы
    ws.append(["РАСХОДЫ"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.append(["Дата", "На что", "Категория", "Комментарий", "Сумма, ₽"])
    _style_header_row(ws, ws.max_row, 5)
    total_exp = 0
    for e in expenses:
        ws.append([e.date, e.name, CAT_LABELS.get(e.category, e.category), e.desc, e.amount])
        total_exp += e.amount
    ws.append(["", "", "", "ИТОГО", total_exp])
    ws.cell(row=ws.max_row, column=4).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=5).font = Font(bold=True)
    ws.append([])
    ws.append(["", "", "", "Прибыль", total_inc - total_exp])
    ws.cell(row=ws.max_row, column=4).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=5).font = Font(bold=True, color="16A34A")

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 22

    # --- Клиенты ---
    wc = wb.create_sheet("Клиенты")
    wc.append(["Имя", "Формат", "Статус", "Старт", "Финиш", "Сумма договора", "Задачи", "Чат", "Холст", "Заметки"])
    _style_header_row(wc, 1, 10)
    for c in clients:
        wc.append([c.name, TYPE_LABELS.get(c.type, c.type), STATUS_LABELS.get(c.status, c.status),
                   c.start, c.end, c.amount, c.tasks, c.chat, c.canvas, c.notes])
    for col in range(1, 11):
        wc.column_dimensions[get_column_letter(col)].width = 20

    # --- Нейроспринт ---
    wn = wb.create_sheet("Нейроспринт")
    wn.append(["Поток", "Старт", "Тариф 1 (кол-во)", "Тариф 2 VIP (кол-во)",
               "Цена Т1", "Цена Т2", "Выручка", "Менеджеры", "Заметки"])
    _style_header_row(wn, 1, 9)
    for s in streams:
        revenue = s.t1 * s.p1 + s.t2 * s.p2
        wn.append([s.name, s.start, s.t1, s.t2, s.p1, s.p2, revenue, s.managers, s.notes])
    for col in range(1, 10):
        wn.column_dimensions[get_column_letter(col)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_csv_finance(incomes, expenses) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["ТИП", "Дата", "Источник/На что", "Категория/Тег", "Описание", "Сумма"])
    for i in incomes:
        w.writerow(["Доход", i.date, i.source, TAG_LABELS.get(i.tag, i.tag), i.desc, i.amount])
    for e in expenses:
        w.writerow(["Расход", e.date, e.name, CAT_LABELS.get(e.category, e.category), e.desc, e.amount])
    return buf.getvalue().encode("utf-8-sig")
